'''assignTask.py:
1 Takes tasks from excel spreadsheet (every task in separate row, starting from row 2)
2 Assignes tasks randomly to all members in database
3 Lets you add/remove/check members in database "teamMembers"
4 Sends emails with assigned tasks to all members. Prints summary.
'''
import openpyxl, shelve, random, smtplib
from openpyxl.utils.cell import get_column_letter

def getTasks(excelFile):
    # getting list of tasks from spreadsheet
    wb = openpyxl.load_workbook(excelFile)
    sheet = wb.active
    lastCell = get_column_letter(sheet.max_column) + str(sheet.max_row)

    tasks = []
    for row in sheet['A2':lastCell]:
        line = []
        for cell in row:
            line.append(cell.value)
        tasks.append(': '.join(line))
    return tasks

def addTeamMember(name, email):
    teamMembersShelf = shelve.open('teamMembers')
    teamMembersShelf[name] = {'email': email}
    teamMembersShelf.close()

def removeTeamMember(name):
    teamMembersShelf = shelve.open('teamMembers')
    del teamMembersShelf[name]
    teamMembersShelf.close()

def showAllCurrentMembers():
    teamMembersShelf = shelve.open('teamMembers')
    for member in teamMembersShelf.keys():
        print('Member: %s, email: %s' % (member, teamMembersShelf[member]['email']))

def createTempDict():
    teamMembersShelf = shelve.open('teamMembers')
    teamMembers = {}
    for key in list(teamMembersShelf.keys()):
        teamMembers.setdefault(key, {})
        teamMembers[key]['email'] = teamMembersShelf[key]['email']
    return teamMembers

def assignTask(tasks, teamMembers):
    for member in teamMembers:
        teamMembers[member]['task'] = []

    # assigning equal numbers of tasks to every member
    while len(tasks) >= len(teamMembers):
        for member in teamMembers:
            randomTask = random.choice(tasks)
            teamMembers[member]['task'].append(randomTask)
            tasks.remove(randomTask)

    # randomly assigning remaining tasks
    while len(tasks) > 0:
        allMembers = list(teamMembers.keys())
        randomMember = random.choice(allMembers)
        randomTask = random.choice(tasks)
        teamMembers[randomMember]['task'].append(randomTask)
        tasks.remove(randomTask)
        allMembers.remove(randomMember)
    
    return teamMembers

def sendEmails(teamMembers):
    try:
        smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
        smtpObj.ehlo()
        smtpObj.starttls()
        print('Enter email:')
        fromEmail = input()
        print('Enter password:')
        password = input()
        smtpObj.login(fromEmail, password)
    except:
        print("Couldn't connect. Check email and password.")

    for member in teamMembers:
        memberTasks = '; '.join(teamMembers[member]['task'])
        email = teamMembers[member]['email']
        message = 'Subject: Extraoppgaver\nHei %s\n\nDenne uka har du fått tildelt følgende ekstraoppgaver:\n%s.\nHilsen' % (member, memberTasks)
        sendMailStatus = smtpObj.sendmail(fromEmail, email, message.encode("utf8"))
        if sendMailStatus != {}:
            print('There was a problem sending email to %s: %s' % (email, sendMailStatus))
        else: 
            print('Email sendt to %s' % member)

    smtpObj.quit()

def printSummary(teamMembers):
    summary = ''
    for member in teamMembers:
        memberTasks = ';\n'.join(teamMembers[member]['task'])
        summary += '%s fikk følgende oppgaver:\n%s\n' % (member, memberTasks)
    return summary

tasks = getTasks('assignTask.xlsx')
teamMembers = createTempDict()
assignTask(tasks, teamMembers)
print(sendEmails(teamMembers))
print(printSummary(teamMembers))