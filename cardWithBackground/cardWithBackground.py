import openpyxl, os
from PIL import Image, ImageDraw, ImageFont
from openpyxl.utils.cell import get_column_letter

def resizeImage(image, modifiedWidth, transparency):
    '''Changes proportions of the image. First argument: image file, second argument: wanted width, third argument: wanted transparency of background image. '''
    #opening given image
    backgroundIm = Image.open(image)
    width, height = backgroundIm.size
    #changing widh to given width, changing height proportionally to given width
    height = int((modifiedWidth / width) * height)
    width = modifiedWidth
    #resizing and saving image
    resizedBackgroundIm = backgroundIm.resize((width, height))
    resizedBackgroundIm.putalpha(transparency)
    return resizedBackgroundIm

def excelToList(excelFile):
    '''Saves text from excel sheet into list of lists'''
    #opens excel sheet
    wb = openpyxl.load_workbook(excelFile)
    sheet = wb.active
    lastCell = get_column_letter(sheet.max_column) + str(sheet.max_row)
    #adding rows to list
    allRowsInList = []
    for rows in (sheet['A2':lastCell]):
        rowInList = []
        for i in range(sheet.max_column):
            rowInList.append(rows[i].value)
        allRowsInList.append(rowInList)
    return allRowsInList

def defineFont(pathToFont, fontSize):
    '''specyfying font that will be used on card'''
    font = ImageFont.truetype(pathToFont, fontSize)
    return font

def checkPlacementOfTextbox(textInList, resizedBackgroundIm, font):
    '''takes a list of lines as first argument, resizedBackgroundIm as second argument (from function "resizeImage") and font from function "definde font"'''
    backgroundWidth, backgroundHeight = resizedBackgroundIm.size
    image = resizedBackgroundIm.copy()
    draw = ImageDraw.Draw(image)
    card = '\n'.join(textInList)
    textWidth, textHeight = draw.textsize(card, font=font, spacing=6)
    coordX = int((backgroundWidth - textWidth) / 2)
    coordY = int((backgroundHeight - textHeight) / 2)
    return image, draw, card, coordX, coordY

def addTextOnImage(draw, card, coordX, coordY):
    '''Adding text on resized image using coordinates, image and text from function checkPlacementOfTextbox'''
    draw.text((coordX, coordY), card, fill='black', font=font, spacing=6)

def makeCards(allRowsInList):
    '''cerating new folder and going thorough list of lists from function "excelToList". Saving images'''
    #new folder for businesscards
    os.makedirs('./businessCards', exist_ok=True)
    #going through list and saving images
    for row in allRowsInList:
        image, draw, card, coordX, coordY = checkPlacementOfTextbox(row, resizedBackgroundIm, font)
        addTextOnImage(draw, card, coordX, coordY)
        image.save(os.path.join('./businessCards', 'businessCard_%s.png' % row[0]))
    print('Folder with all cards created.')

def addFrame(resizedBackgroundIm):
    '''Adding frame to image'''
    draw = ImageDraw.Draw(resizedBackgroundIm)
    width, height = resizedBackgroundIm.size
    draw.line([(0, 0), (width - 1, 0), (width - 1, height - 1), (0, height - 1), (0, 0)], fill='black')

def noOfImOnPageForPrint(resizedBackgroundIm):
    '''calculating how many images will fit in one A4 page'''
    width, height = resizedBackgroundIm.size
    maxNoOfImVertical = 1000 // width
    maxNoOfImHorizontal = 700 // height
    maxNoOfIm = maxNoOfImVertical * maxNoOfImHorizontal
    pageForPrintWidth = maxNoOfImHorizontal * width
    pageForPrintHeight = maxNoOfImVertical * height
    return maxNoOfIm, pageForPrintWidth, pageForPrintHeight, maxNoOfImHorizontal


def prepareForPrint(folder, maxNoOfIm, pageForPrintWidth, pageForPrintHeight, maxNoOfImHorizontal, resizedBackgroundIm):
    '''saving images on pages for print'''
    listOfImages = os.listdir(folder)
    width, height = resizedBackgroundIm.size
    # calculating margins:
    marginTop = (700 - pageForPrintHeight) // 2
    marginLeft = (1000 - pageForPrintWidth) // 2

    imLeft = len(listOfImages) % maxNoOfIm
    pageForPrintNo = 1
    # creating full pages:
    while len(listOfImages) > imLeft:
        im = Image.new('RGBA', (1000, 700), 'white')
        imNo = 1
        while imNo <= maxNoOfIm:
            for top in range(marginTop, pageForPrintHeight, height):
                for left in range(marginLeft, pageForPrintWidth, width):
                    fileToPaste = Image.open(os.path.join(folder, listOfImages[0]))
                    addFrame(fileToPaste)
                    im.paste(fileToPaste, (left, top))
                    listOfImages.pop(0)
                    imNo += 1
            im.save('forPrint%s.png' % pageForPrintNo)
            pageForPrintNo += 1
    # creating last page (if some leftovers after cerating full pages):
    else:
        im = Image.new('RGBA', (1000, 700), 'white')
        wholeLines = imLeft // maxNoOfImHorizontal
        imInLastLine = 0
        if imLeft % maxNoOfImHorizontal != 0:
            imInLastLine = imLeft % maxNoOfImHorizontal
        for top in range(marginTop, wholeLines * height, height):
            for left in range(marginLeft, maxNoOfImHorizontal * width, width):
                fileToPaste = Image.open(os.path.join(folder, listOfImages[0]))
                addFrame(fileToPaste)
                im.paste(fileToPaste, (left, top))
                listOfImages.pop(0)
        if imInLastLine != 0:
            for i in range(imInLastLine):
                fileToPaste = Image.open(os.path.join(folder, listOfImages[0]))
                addFrame(fileToPaste)
                im.paste(fileToPaste, (marginLeft + (i * width), marginTop + (wholeLines * height)))
                listOfImages.pop(0)
        im.save('forPrint%s.png' % pageForPrintNo)
        print('Pages for print created and saved.')


image = 'paper.jpg'
modifiedWidth = 220
transparency = 150
resizedBackgroundIm = resizeImage(image, modifiedWidth, transparency)

excelFile = 'cardWithBackground.xlsx'
allRowsInList = excelToList(excelFile)

pathToFont = 'C:\Windows\Fonts\calibri\calibri.ttf'
fontSize = 18
font = defineFont(pathToFont, fontSize)

makeCards(allRowsInList)

maxNoOfIm, pageForPrintWidth, pageForPrintHeight, maxNoOfImHorizontal = noOfImOnPageForPrint(resizedBackgroundIm)
prepareForPrint('./businessCards', maxNoOfIm, pageForPrintWidth, pageForPrintHeight, maxNoOfImHorizontal, resizedBackgroundIm)


